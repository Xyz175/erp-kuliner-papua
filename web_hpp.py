import streamlit as st
import pandas as pd
import psycopg2
import hashlib
import io
import base64
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 0. TEMA WARNA ETNIK PAPUA (DARK MODE)
# ==========================================
C_PRIMARY = "#C57A2E"       # Oranye khas logo
C_PRIMARY_DARK = "#8F5620"
LOGO_FILE = "logo_rumah_etnik_papua.png"  

# ==========================================
# 1. KONFIGURASI HALAMAN & DATABASE CLOUD
# ==========================================
st.set_page_config(
    page_title="ERP Rumah Etnik Papua",
    layout="wide",
    page_icon=LOGO_FILE if __import__("os").path.exists(LOGO_FILE) else "🏢"
)

# --- KREDENSIAL DATABASE ---
_FALLBACK_URI = "postgresql://postgres.jskxygpvnbjgjjgsvwqf:ZeeyStore175@aws-0-ap-southeast-2.pooler.supabase.com:6543/postgres"
SUPABASE_URI = st.secrets.get("SUPABASE_URI", _FALLBACK_URI) if hasattr(st, "secrets") else _FALLBACK_URI

def get_connection():
    return psycopg2.connect(SUPABASE_URI)

def hash_password(raw_password: str) -> str:
    return hashlib.sha256(raw_password.strip().encode("utf-8")).hexdigest()

def init_db():
    conn = get_connection()
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute("CREATE TABLE IF NOT EXISTS users (username TEXT PRIMARY KEY, password TEXT, role TEXT)")
    cursor.execute(
        "INSERT INTO users (username, password, role) VALUES ('owner', %s, 'Owner') ON CONFLICT DO NOTHING",
        (hash_password("owner123"),)
    )

    # Migrasi keamanan password
    cursor.execute("SELECT username, password FROM users")
    for uname, pw in cursor.fetchall():
        if pw is None or len(pw) != 64:
            cursor.execute("UPDATE users SET password=%s WHERE username=%s", (hash_password(pw or ""), uname))

    # Tabel Gudang Pusat & Cafe
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bahan_baku (
            nama_bahan TEXT PRIMARY KEY, harga_beli REAL, jumlah_isi REAL, satuan TEXT, harga_satuan REAL,
            stok_gudang REAL DEFAULT 0, stok_cafe REAL DEFAULT 0, stok_taman REAL DEFAULT 0, harga_jual_internal REAL DEFAULT 0
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS resep (
            id SERIAL PRIMARY KEY, nama_resep TEXT UNIQUE, total_hpp_bahan REAL,
            total_operasional REAL, total_hpp_final REAL, harga_jual REAL DEFAULT 0, tanggal_dibuat TEXT
        )
    """)
    cursor.execute("CREATE TABLE IF NOT EXISTS detail_resep (id SERIAL PRIMARY KEY, resep_id INTEGER, nama_bahan TEXT, jumlah_pakai REAL, satuan TEXT, subtotal_biaya REAL)")
    cursor.execute("CREATE TABLE IF NOT EXISTS riwayat_stok (id SERIAL PRIMARY KEY, tanggal TEXT, nama_bahan TEXT, jenis_transaksi TEXT, jumlah REAL, keterangan TEXT, laba_internal REAL DEFAULT 0, dari_gudang TEXT, ke_unit TEXT)")
    
    # 🌟 PEMBARUAN TABEL GUDANG KIOS 🌟
    cursor.execute("CREATE TABLE IF NOT EXISTS stok_kios_item (nama_barang TEXT PRIMARY KEY, stok REAL DEFAULT 0, satuan TEXT)")
    
    # Mengamankan database dengan menambahkan kolom baru (jika belum ada)
    try: cursor.execute("ALTER TABLE stok_kios_item ADD COLUMN harga_beli REAL DEFAULT 0")
    except: pass
    try: cursor.execute("ALTER TABLE stok_kios_item ADD COLUMN harga_jual REAL DEFAULT 0")
    except: pass
    try: cursor.execute("ALTER TABLE stok_kios_item ADD COLUMN stok_etalase REAL DEFAULT 0")
    except: pass

    cursor.execute("CREATE TABLE IF NOT EXISTS riwayat_kios (id SERIAL PRIMARY KEY, tanggal TEXT, nama_barang TEXT, jenis_transaksi TEXT, jumlah REAL, pic TEXT)")
    
    # Tabel Laporan Validasi Kios Harian
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS laporan_penjualan_kios (
            id SERIAL PRIMARY KEY, tanggal TEXT, nama_barang TEXT,
            stok_awal_sistem REAL, terjual REAL, dihutang REAL, hilang REAL, sisa_fisik_nyata REAL,
            laba_rp REAL, rugi_rp REAL, pic TEXT
        )
    """)

    cursor.close()
    conn.close()

try:
    init_db()
except Exception as e:
    st.error(f"⚠️ Gagal sinkronisasi arsitektur DB: {e}")

# ==========================================
# 1B. HELPER EXPORT EXCEL 
# ==========================================
def buat_excel_menarik(df: pd.DataFrame, title: str, sheet_name: str = "Data", currency_cols=None, extra_info=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(df.columns)
    last_col = get_column_letter(len(headers))
    border = Border(left=Side(style="thin", color="D8CBB8"), right=Side(style="thin", color="D8CBB8"), 
                    top=Side(style="thin", color="D8CBB8"), bottom=Side(style="thin", color="D8CBB8"))
    currency_cols = currency_cols or []

    row_cursor = 1
    ws.merge_cells(f"A{row_cursor}:{last_col}{row_cursor}")
    c = ws.cell(row=row_cursor, column=1, value=f"  {title}")
    c.font = Font(size=15, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="C57A2E", end_color="C57A2E", fill_type="solid")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_cursor].height = 30
    row_cursor += 1

    extra_info = extra_info or [f"Diekspor pada: {datetime.now().strftime('%d %B %Y, %H:%M')}"]
    for info in extra_info:
        ws.merge_cells(f"A{row_cursor}:{last_col}{row_cursor}")
        ic = ws.cell(row=row_cursor, column=1, value=f"  {info}")
        ic.font = Font(italic=True, size=9, color="7A6A57")
        row_cursor += 1
    row_cursor += 1

    header_row = row_cursor
    for col_idx, h in enumerate(headers, start=1):
        hc = ws.cell(row=header_row, column=col_idx, value=str(h))
        hc.font = Font(bold=True, color="FFFFFF", size=10)
        hc.fill = PatternFill(start_color="D98C3D", end_color="D98C3D", fill_type="solid")
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border = border
    row_cursor += 1

    col_max_len = [len(str(h)) for h in headers]
    for r_off, (_, row) in enumerate(df.iterrows()):
        r_idx = row_cursor + r_off
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if r_off % 2 == 1:
                cell.fill = PatternFill(start_color="FBF0E4", end_color="FBF0E4", fill_type="solid")
            if (c_idx - 1) in currency_cols and isinstance(val, (int, float)):
                cell.number_format = '"Rp" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            col_max_len[c_idx - 1] = max(col_max_len[c_idx - 1], len(str(val)))

    for col_idx, ml in enumerate(col_max_len, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(ml + 4, 12), 45)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def tombol_download_excel(df: pd.DataFrame, label: str, file_name: str, title: str, currency_cols=None, key=None):
    if df is None or df.empty: return
    data = buat_excel_menarik(df, title=title, currency_cols=currency_cols)
    st.download_button(label, data=data, file_name=file_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key=key)

def get_daftar_bahan():
    try:
        conn = get_connection()
        df = pd.read_sql_query("SELECT * FROM bahan_baku ORDER BY nama_bahan ASC", conn)
        conn.close()
        return df
    except: return pd.DataFrame()

if 'racikan_sementara' not in st.session_state:
    st.session_state.racikan_sementara = pd.DataFrame(columns=["Nama Bahan", "Jumlah Pakai", "Satuan", "Subtotal"])
if 'sudah_login' not in st.session_state:
    st.session_state.sudah_login = False

# ==========================================
# 1C. GAYA VISUAL (CSS TEMA GELAP ELEGAN)
# ==========================================
st.markdown(f"""
<style>
    .stApp {{ background-color: #121212; }}
    section[data-testid="stSidebar"] {{ background-color: #1A1A1A; border-right: 2px solid {C_PRIMARY_DARK}; }}
    p, span, label, .stMarkdown {{ color: #D4C4A8 !important; }}
    h1, h2, h3, h4 {{ color: {C_PRIMARY} !important; }}
    .stTextInput>div>div>input, .stNumberInput>div>div>input {{ background-color: #262626 !important; color: #D4C4A8 !important; border: 1px solid {C_PRIMARY_DARK} !important; }}
    div[data-testid="stMetric"] {{ background-color: #1E1E1E; border: 1px solid {C_PRIMARY_DARK}; border-radius: 10px; padding: 10px; }}
    .stButton>button, .stDownloadButton>button {{ background-color: {C_PRIMARY}; color: #121212 !important; border: none; border-radius: 8px; font-weight: 700; }}
    .stButton>button:hover, .stDownloadButton>button:hover {{ background-color: {C_PRIMARY_DARK}; color: #D4C4A8 !important; }}
    .footer-zeey {{ margin-top: 60px; padding: 22px 10px; text-align: center; border-top: 1px dashed {C_PRIMARY_DARK}; color: #8F5620; font-size: 14px; }}
    .footer-zeey b {{ color: {C_PRIMARY}; }}
</style>
""", unsafe_allow_html=True)

def render_footer():
    st.markdown(f"""
    <div class="footer-zeey">
        Dibuat &amp; dikembangkan oleh <b>ZeeyStudio</b><br>
        📱 WhatsApp: <b>0813-5413-1178</b>
    </div>
    """, unsafe_allow_html=True)

# ==========================================
# 2. GERBANG UTAMA (SISTEM LOGIN)
# ==========================================
if not st.session_state.sudah_login:
    import os
    _, col_center, _ = st.columns([1, 1.2, 1])

    with col_center:
        st.write("<br><br>", unsafe_allow_html=True) 
        if os.path.exists(LOGO_FILE):
            st.image(LOGO_FILE, use_container_width=True)
            
        st.write("<br>", unsafe_allow_html=True)
        st.markdown(f"""<div style="background-color: #1E1E1E; border: 1px solid {C_PRIMARY_DARK}; border-radius: 16px; padding: 30px;"><h3 style="text-align: center; margin-bottom: 20px;">🔐 Masuk ke Sistem</h3>""", unsafe_allow_html=True)
        
        user_input = st.text_input("Username")
        pass_input = st.text_input("Password", type="password")
        
        st.write("<br>", unsafe_allow_html=True)
        if st.button("Masuk", type="primary", use_container_width=True):
            try:
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT role FROM users WHERE username=%s AND password=%s", (user_input.strip().lower(), hash_password(pass_input)))
                res = cursor.fetchone()
                conn.close()
                if res:
                    st.session_state.sudah_login = True
                    st.session_state.username = user_input.strip().lower()
                    st.session_state.role = res[0]
                    st.rerun()
                else: st.error("❌ Username atau Password salah!")
            except Exception as e: st.error(f"Gagal terhubung ke database: {e}")
        st.markdown("</div>", unsafe_allow_html=True)
    render_footer()
    st.stop()

# ==========================================
# 3. NAVIGASI BAR BERDASARKAN ROLE
# ==========================================
import os
role = st.session_state.role

if os.path.exists(LOGO_FILE):
    st.sidebar.image(LOGO_FILE, use_container_width=True)
    st.sidebar.write("---")

st.sidebar.write(f"👤 User: **{st.session_state.username}** ({role})")
if st.sidebar.button("🚪 Logout", use_container_width=True):
    st.session_state.sudah_login = False
    st.rerun()
st.sidebar.write("---")

if role == "Gudang Kios":
    menu_options = ["🏪 Operasional Kios"]
elif role == "Karyawan":
    menu_options = ["🍽️ HPP Cafe"]
elif role == "Admin":
    menu_options = ["📊 Dashboard Admin", "📦 Gudang Induk", "🏪 Operasional Kios", "🍽️ HPP Cafe"]
elif role == "Owner":
    menu_options = ["📊 Dashboard Admin", "📦 Gudang Induk", "🏪 Operasional Kios", "🍽️ HPP Cafe", "👥 Menu Owner"]
else:
    menu_options = []

menu = st.sidebar.radio("Menu Navigasi:", menu_options) if menu_options else None

# ==========================================
# 4. IMPLEMENTASI OPERASIONAL
# ==========================================

# --- MENU 1: DASHBOARD ADMIN ---
if menu == "📊 Dashboard Admin":
    st.title("📊 Dashboard Analitik Finansial")
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT SUM(laba_internal) FROM riwayat_stok WHERE jenis_transaksi='Penjualan Internal'")
        total_laba_gudang = cursor.fetchone()[0] or 0.0

        cursor.execute("SELECT SUM(harga_jual - total_hpp_final) FROM resep WHERE harga_jual > 0")
        total_laba_cafe = cursor.fetchone()[0] or 0.0
        
        cursor.execute("SELECT SUM(laba_rp), SUM(rugi_rp) FROM laporan_penjualan_kios")
        laba_kios, rugi_kios = cursor.fetchone()
        laba_kios = laba_kios or 0.0
        rugi_kios = rugi_kios or 0.0

        m1, m2, m3 = st.columns(3)
        m1.metric("💰 Laba Internal Gudang Induk", f"Rp {total_laba_gudang:,.0f}")
        m2.metric("🍽️ Estimasi Laba Resep Cafe", f"Rp {total_laba_cafe:,.0f}")
        m3.metric("🏪 Laba Bersih Kios (-Rugi)", f"Rp {(laba_kios - rugi_kios):,.0f}", f"Rugi Fisik: Rp {rugi_kios:,.0f}")

        st.write("---")
        st.subheader("📜 Riwayat Laba Rugi Kios (Validasi)")
        df_log_k = pd.read_sql_query("SELECT tanggal as \"Waktu\", nama_barang as \"Barang\", terjual as \"Laku\", dihutang as \"Bon\", hilang as \"Rugi Fisik\", laba_rp as \"Laba (Rp)\", rugi_rp as \"Kerugian (Rp)\", pic as \"Checker\" FROM laporan_penjualan_kios ORDER BY id DESC", conn)
        st.dataframe(df_log_k, use_container_width=True)
        tombol_download_excel(df_log_k, "📥 Unduh Laporan Kios (Excel)", "LabaRugi_Kios.xlsx", title="Laporan Profit & Loss Kios", currency_cols=[5, 6])
        conn.close()
    except Exception as e:
        st.warning(f"Menunggu sinkronisasi data... ({e})")

# --- MENU 2: GUDANG INDUK ---
elif menu == "📦 Gudang Induk":
    st.title("📦 Pusat Logistik & Gudang Utama")

    tab_induk, tab_transfer = st.tabs(["📥 Stok Induk Pusat", "🚚 Transfer Lintas Cabang"])
    
    with tab_induk:
        df_b = get_daftar_bahan()
        with st.form("form_pusat", clear_on_submit=True):
            cc1, cc2, cc3 = st.columns(3)
            with cc1: nama = st.text_input("Nama Komoditas Baru:")
            with cc2: hrg = st.number_input("Harga Beli Total (Rp):", min_value=0.0)
            with cc3: qty = st.number_input("Total Isi/Volume:", min_value=1.0)
            hrg_int = st.number_input("Harga Jual Internal (ke Cafe/Taman):", min_value=0.0)
            sat = st.text_input("Satuan (gram, pcs, ml):")
            if st.form_submit_button("💾 Masukkan ke Brankas Induk"):
                if nama.strip() and sat.strip():
                    conn = get_connection()
                    cursor = conn.cursor()
                    harga_per_satuan = float(hrg / qty)
                    cursor.execute(
                        "INSERT INTO bahan_baku (nama_bahan, harga_beli, jumlah_isi, satuan, harga_satuan, stok_gudang, harga_jual_internal) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (nama_bahan) DO UPDATE SET stok_gudang = bahan_baku.stok_gudang + EXCLUDED.stok_gudang, harga_jual_internal=EXCLUDED.harga_jual_internal",
                        (nama.strip(), float(hrg), float(qty), sat.strip(), harga_per_satuan, float(qty), float(hrg_int))
                    )
                    conn.commit()
                    conn.close()
                    st.toast(f"✅ {nama} Tersimpan di Induk!", icon="📥")
                    st.rerun()
                else: st.warning("Nama dan satuan wajib diisi.")
        st.write("---")
        if not df_b.empty:
            df_tampil = df_b[['nama_bahan', 'harga_satuan', 'harga_jual_internal', 'stok_gudang', 'stok_cafe', 'stok_taman', 'satuan']].copy()
            df_tampil.columns = ['Nama Barang', 'Modal Dasar', 'Jual Internal', 'STOK INDUK', 'STOK CAFE', 'STOK TAMAN', 'Satuan']
            st.dataframe(df_tampil, use_container_width=True)

    with tab_transfer:
        df_g = get_daftar_bahan()
        if df_g.empty: st.warning("Stok Gudang Induk kosong.")
        else:
            c1, c2, c3 = st.columns(3)
            with c1: brg = st.selectbox("Pilih Material:", df_g['nama_bahan'].tolist())
            data_b = df_g[df_g['nama_bahan'] == brg].iloc[0]
            stok_induk = data_b['stok_gudang']
            with c2: qty_tr = st.number_input(f"Kirim (Tersedia: {stok_induk}):", min_value=0.0, max_value=float(stok_induk) if stok_induk > 0 else 0.0)
            with c3: tuj = st.radio("Kirim Ke Cabang:", ["Cafe", "Taman"])
            if st.button("🚀 Kirim Barang", type="primary", use_container_width=True):
                if qty_tr <= 0: st.warning("Harus lebih dari 0.")
                elif stok_induk >= qty_tr:
                    tgl = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    harga_acuan = data_b['harga_jual_internal'] if data_b['harga_jual_internal'] > 0 else data_b['harga_satuan']
                    laba = float((qty_tr * harga_acuan) - (qty_tr * data_b['harga_satuan']))
                    conn = get_connection()
                    cursor = conn.cursor()
                    cursor.execute("UPDATE bahan_baku SET stok_gudang = stok_gudang - %s WHERE nama_bahan = %s", (float(qty_tr), brg))
                    if tuj == "Cafe": cursor.execute("UPDATE bahan_baku SET stok_cafe = COALESCE(stok_cafe, 0) + %s WHERE nama_bahan = %s", (float(qty_tr), brg))
                    elif tuj == "Taman": cursor.execute("UPDATE bahan_baku SET stok_taman = COALESCE(stok_taman, 0) + %s WHERE nama_bahan = %s", (float(qty_tr), brg))
                    cursor.execute("INSERT INTO riwayat_stok (tanggal, nama_bahan, jenis_transaksi, jumlah, keterangan, laba_internal, dari_gudang, ke_unit) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)", (tgl, brg, 'Penjualan Internal', float(qty_tr), str(qty_tr * harga_acuan), laba, 'Gudang Etnik', tuj))
                    conn.commit()
                    conn.close()
                    st.toast(f"🚀 Dikirim ke {tuj}!", icon="🚚")
                    st.rerun()

# --- MENU 2B: OPERASIONAL KIOS (GUDANG KIOS & OWNER) ---
elif menu == "🏪 Operasional Kios":
    st.title("🏪 Tata Kelola Gudang & Etalase Kios")
    
    conn = get_connection()
    df_kios = pd.read_sql_query("SELECT * FROM stok_kios_item ORDER BY nama_barang ASC", conn)

    # Susunan Tab menyesuaikan Role
    if role in ["Owner", "Admin"]:
        tab_master, tab_masuk, tab_keluar, tab_validasi, tab_rep = st.tabs(["⚙️ Master Kios (Pusat)", "📥 Masuk Gudang Kios", "📤 Keluar ke Etalase", "✅ Validasi Harian", "📊 Riwayat Logistik"])
    else:
        tab_masuk, tab_keluar, tab_validasi, tab_rep = st.tabs(["📥 Masuk Gudang Kios", "📤 Keluar ke Etalase", "✅ Validasi Harian", "📊 Riwayat Logistik"])
        tab_master = None

    # HANYA OWNER & ADMIN yang bisa menambah barang Kios
    if tab_master:
        with tab_master:
            st.info("💡 Hanya Admin dan Owner yang bisa mendaftarkan barang, mengatur Harga Modal (Beli), dan Harga Jual Kios. Karyawan Gudang hanya bisa menambah stoknya.")
            with st.form("form_master_kios", clear_on_submit=True):
                k1, k2, k3, k4 = st.columns(4)
                with k1: m_nama = st.text_input("Nama Barang Kios:")
                with k2: m_beli = st.number_input("Harga Beli / Modal (Rp):", min_value=0.0)
                with k3: m_jual = st.number_input("Harga Jual Konsumen (Rp):", min_value=0.0)
                with k4: m_sat = st.text_input("Satuan (pcs, bks, dll):", value="pcs")
                if st.form_submit_button("💾 Daftarkan Item Kios", use_container_width=True):
                    if m_nama.strip():
                        cursor = conn.cursor()
                        cursor.execute("""
                            INSERT INTO stok_kios_item (nama_barang, harga_beli, harga_jual, satuan, stok, stok_etalase)
                            VALUES (%s, %s, %s, %s, 0, 0)
                            ON CONFLICT (nama_barang) DO UPDATE SET 
                            harga_beli=EXCLUDED.harga_beli, harga_jual=EXCLUDED.harga_jual, satuan=EXCLUDED.satuan
                        """, (m_nama.strip(), float(m_beli), float(m_jual), m_sat.strip()))
                        conn.commit()
                        st.toast(f"✅ Master {m_nama} tersimpan!", icon="⚙️")
                        st.rerun()

    with tab_masuk:
        st.subheader("📥 Barang Masuk Gudang Belakang (Kios)")
        if df_kios.empty: st.warning("Belum ada daftar barang. Hubungi Owner/Admin untuk mendaftarkan Master Kios.")
        else:
            with st.form("form_masuk_kios", clear_on_submit=True):
                c_in1, c_in2 = st.columns(2)
                with c_in1: b_masuk = st.selectbox("Pilih Barang:", df_kios['nama_barang'].tolist())
                with c_in2: q_masuk = st.number_input("Kuantitas Masuk Gudang:", min_value=1.0, step=1.0)
                if st.form_submit_button("📥 Tambah ke Gudang Kios"):
                    cursor = conn.cursor()
                    tgl = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    cursor.execute("UPDATE stok_kios_item SET stok = stok + %s WHERE nama_barang = %s", (float(q_masuk), b_masuk))
                    cursor.execute("INSERT INTO riwayat_kios (tanggal, nama_barang, jenis_transaksi, jumlah, pic) VALUES (%s, %s, %s, %s, %s)", (tgl, b_masuk, 'Masuk Gudang', float(q_masuk), st.session_state.username))
                    conn.commit()
                    st.toast(f"📦 {q_masuk} {b_masuk} masuk ke Gudang Kios!", icon="✅")
                    st.rerun()

    with tab_keluar:
        st.subheader("📤 Pindahkan Barang dari Gudang ke Etalase (Toko)")
        if df_kios.empty: st.warning("Kosong.")
        else:
            with st.form("form_keluar_kios", clear_on_submit=True):
                c_out1, c_out2 = st.columns(2)
                with c_out1: b_keluar = st.selectbox("Pilih Barang Gudang:", df_kios['nama_barang'].tolist())
                sisa_gdg = df_kios[df_kios['nama_barang'] == b_keluar]['stok'].values[0] if not df_kios.empty else 0
                with c_out2: q_keluar = st.number_input(f"Keluarkan ke Rak/Etalase (Sisa Gudang: {sisa_gdg}):", min_value=0.0, max_value=float(sisa_gdg) if sisa_gdg > 0 else 0.0, step=1.0)
                
                if st.form_submit_button("📤 Taruh di Etalase Toko"):
                    if q_keluar > 0:
                        cursor = conn.cursor()
                        tgl = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        cursor.execute("UPDATE stok_kios_item SET stok = stok - %s, stok_etalase = stok_etalase + %s WHERE nama_barang = %s", (float(q_keluar), float(q_keluar), b_keluar))
                        cursor.execute("INSERT INTO riwayat_kios (tanggal, nama_barang, jenis_transaksi, jumlah, pic) VALUES (%s, %s, %s, %s, %s)", (tgl, b_keluar, 'Pindah ke Etalase', float(q_keluar), st.session_state.username))
                        conn.commit()
                        st.toast(f"🛒 {b_keluar} dipajang di Etalase!", icon="✅")
                        st.rerun()

    with tab_validasi:
        st.subheader("✅ Validasi Stok Penjualan Harian (Stock Opname)")
        st.info("Cocokkan sisa fisik barang yang ada di etalase/rak toko. Sistem akan otomatis menghitung laba dan kerugiannya.")
        
        # Hanya tampilkan barang yang ada di etalase
        df_etalase = df_kios[df_kios['stok_etalase'] > 0]
        if df_etalase.empty: st.warning("Tidak ada barang di etalase Kios.")
        else:
            with st.form("form_validasi_kios"):
                v_brg = st.selectbox("Barang Etalase yang Divalidasi:", df_etalase['nama_barang'].tolist())
                row_v = df_etalase[df_etalase['nama_barang'] == v_brg].iloc[0]
                stok_sys = float(row_v['stok_etalase'])
                h_beli = float(row_v['harga_beli'])
                h_jual = float(row_v['harga_jual'])

                st.markdown(f"**Stok Sistem Etalase Saat Ini:** `{stok_sys}` {row_v['satuan']}")
                
                v1, v2, v3, v4 = st.columns(4)
                with v1: laku = st.number_input("Jumlah Laku Dibeli:", min_value=0.0, step=1.0)
                with v2: hutang = st.number_input("Diambil / Dihutang:", min_value=0.0, step=1.0)
                with v3: hilang = st.number_input("Hilang / Rusak (Kerugian):", min_value=0.0, step=1.0)
                with v4: sisa_fisik = st.number_input("SISA FISIK NYATA DI TOKO:", min_value=0.0, step=1.0)

                if st.form_submit_button("✅ Kunci Laporan Kios Hari Ini", type="primary", use_container_width=True):
                    # Hitung Finansial
                    laba = float(laku * (h_jual - h_beli))
                    rugi = float(hilang * h_beli)
                    
                    cursor = conn.cursor()
                    tgl = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    # Update stok etalase menjadi sisa fisik nyata yang diinput
                    cursor.execute("UPDATE stok_kios_item SET stok_etalase = %s WHERE nama_barang = %s", (float(sisa_fisik), v_brg))
                    
                    # Simpan ke tabel Laba Rugi Kios
                    cursor.execute("""
                        INSERT INTO laporan_penjualan_kios (tanggal, nama_barang, stok_awal_sistem, terjual, dihutang, hilang, sisa_fisik_nyata, laba_rp, rugi_rp, pic)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (tgl, v_brg, stok_sys, float(laku), float(hutang), float(hilang), float(sisa_fisik), laba, rugi, st.session_state.username))
                    
                    conn.commit()
                    st.toast(f"✅ Laporan {v_brg} terkunci! Laba: Rp {laba:,.0f} | Rugi: Rp {rugi:,.0f}", icon="📊")
                    st.rerun()

    with tab_rep:
        st.subheader("Data Persediaan Kios (Real-Time)")
        st.dataframe(df_kios[['nama_barang', 'harga_beli', 'harga_jual', 'stok', 'stok_etalase', 'satuan']].rename(columns={'stok': 'Stok Gudang', 'stok_etalase': 'Stok Rak Toko'}), use_container_width=True)
        
        st.write("---")
        st.subheader("Log Aktivitas Kios (Mutasi)")
        df_log_kios = pd.read_sql_query("SELECT tanggal as \"Waktu\", nama_barang as \"Barang\", jenis_transaksi as \"Status\", jumlah as \"Qty\", pic as \"Operator\" FROM riwayat_kios ORDER BY id DESC", conn)
        st.dataframe(df_log_kios, use_container_width=True)
    conn.close()

# --- MENU 3: HPP CAFE ---
elif menu == "🍽️ HPP Cafe":
    st.title("🍽️ Pusat Operasional & Manajemen HPP Cafe")

    if role in ["Owner", "Admin"]:
        tab_dapur, tab_tambah_bahan, tab_arsip = st.tabs(["🍳 Input Pemakaian Dapur", "➕ Tambah Bahan Cafe", "📜 Arsip & Harga Jual"])
    else:
        tab_dapur, tab_tambah_bahan = st.tabs(["🍳 Input Pemakaian Dapur", "➕ Tambah Bahan Cafe"])

    with tab_dapur:
        st.subheader("🍳 Meracik Menu & Potong Stok")
        df_b = get_daftar_bahan()
        if df_b.empty: st.warning("Belum ada bahan. Silakan isi di tab 'Tambah Bahan Cafe'.")
        else:
            c1, c2, c3 = st.columns([2, 1, 1])
            with c1:
                cari = st.text_input("🔍 Cari bahan baku Cafe:")
                df_f = df_b[df_b['nama_bahan'].str.lower().str.contains(cari.lower())] if cari else df_b
                pilih_b = st.selectbox("Pilih item:", df_f['nama_bahan'].tolist() if not df_f.empty else ["Kosong"])
            with c2: qty = st.number_input("Volume / Pakai:", min_value=0.0, step=1.0)
            with c3:
                st.write(""); st.write("")
                if st.button("➕ Tambah ke Draf Menu", use_container_width=True):
                    if pilih_b != "Kosong" and qty > 0:
                        row = df_b[df_b['nama_bahan'] == pilih_b].iloc[0]
                        h_modal = row['harga_jual_internal'] if row['harga_jual_internal'] > 0 else row['harga_satuan']
                        sub = float(qty * h_modal)
                        df_skrg = st.session_state.racikan_sementara
                        if pilih_b in df_skrg['Nama Bahan'].values:
                            df_skrg.loc[df_skrg['Nama Bahan'] == pilih_b, ["Jumlah Pakai", "Subtotal"]] = [float(qty), sub]
                        else:
                            tambah = pd.DataFrame([[str(pilih_b), float(qty), str(row['satuan']), sub]], columns=df_skrg.columns)
                            st.session_state.racikan_sementara = pd.concat([df_skrg, tambah], ignore_index=True)
                        st.rerun()
        st.write("---")
        if not st.session_state.racikan_sementara.empty:
            st.session_state.racikan_sementara = st.data_editor(st.session_state.racikan_sementara, use_container_width=True)
            tot_bahan = float(st.session_state.racikan_sementara['Subtotal'].sum())

            st.subheader("⚡ Komponen Operasional Dapur")
            co1, co2 = st.columns(2)
            with co1: waktu = st.number_input("Waktu Meracik (Menit):", min_value=0.0, step=5.0)
            with co2: gas = st.number_input("Gas LPG (Kg):", min_value=0.0, format="%.3f")

            biaya_ops = float(((10000 / 60) * waktu) + ((3000 / 60) * waktu) + (gas * 18000))
            tot_final = float(tot_bahan + biaya_ops)

            st.info(f"📊 Modal Bahan: Rp {tot_bahan:,.0f} | Operasional: Rp {biaya_ops:,.0f} | HPP Pokok: Rp {tot_final:,.2f}")

            st.subheader("💰 Estimasi Harga Jual & Laba")
            hj1, hj2 = st.columns(2)
            with hj1: harga_jual_draf = st.number_input("Rencana Harga Jual Menu (Rp):", min_value=0.0, step=1000.0)
            with hj2:
                laba_draf = harga_jual_draf - tot_final
                margin_draf = (laba_draf / harga_jual_draf * 100) if harga_jual_draf > 0 else 0.0
                st.markdown(f"**Estimasi Laba:** Rp {laba_draf:,.2f} ({margin_draf:.1f}%)")

            cx1, cx2 = st.columns(2)
            with cx1: nama_menu = st.text_input("Nama Menu Kuliner:")
            with cx2:
                st.write(""); st.write("")
                if st.button("💾 Kunci Resep & Potong Stok", type="primary", use_container_width=True):
                    if nama_menu.strip():
                        conn = get_connection()
                        cursor = conn.cursor()
                        tgl = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        cursor.execute(
                            "INSERT INTO resep (nama_resep, total_hpp_bahan, total_operasional, total_hpp_final, harga_jual, tanggal_dibuat) VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (nama_resep) DO UPDATE SET total_hpp_bahan=EXCLUDED.total_hpp_bahan, total_operasional=EXCLUDED.total_operasional, total_hpp_final=EXCLUDED.total_hpp_final, harga_jual=EXCLUDED.harga_jual",
                            (nama_menu.strip(), tot_bahan, biaya_ops, tot_final, harga_jual_draf, tgl)
                        )
                        cursor.execute("SELECT id FROM resep WHERE nama_resep=%s", (nama_menu.strip(),))
                        r_id = cursor.fetchone()[0]
                        cursor.execute("DELETE FROM detail_resep WHERE resep_id=%s", (r_id,))

                        for _, r in st.session_state.racikan_sementara.iterrows():
                            cursor.execute("INSERT INTO detail_resep (resep_id, nama_bahan, jumlah_pakai, satuan, subtotal_biaya) VALUES (%s,%s,%s,%s,%s)", (r_id, str(r['Nama Bahan']), float(r['Jumlah Pakai']), str(r['Satuan']), float(r['Subtotal'])))
                            cursor.execute("UPDATE bahan_baku SET stok_cafe = COALESCE(stok_cafe,0) - %s WHERE nama_bahan = %s", (float(r['Jumlah Pakai']), str(r['Nama Bahan'])))

                        conn.commit()
                        conn.close()
                        st.session_state.racikan_sementara = pd.DataFrame(columns=["Nama Bahan", "Jumlah Pakai", "Satuan", "Subtotal"])
                        st.toast("🎉 Resep Disimpan!", icon="🍳")
                        st.rerun()

    with tab_tambah_bahan:
        with st.form("form_tambah_bahan_cafe", clear_on_submit=True):
            col_b1, col_b2 = st.columns(2)
            with col_b1: nama_baru_cafe = st.text_input("Nama Bahan Baku:")
            with col_b2: hrg_baru_cafe = st.number_input("Harga Beli (Total Rp):", min_value=0.0)
            col_b3, col_b4 = st.columns(2)
            with col_b3: qty_baru_cafe = st.number_input("Jumlah / Isi Berat:", min_value=1.0)
            with col_b4: sat_baru_cafe = st.text_input("Satuan (kg, pcs, ml, dll):")

            if st.form_submit_button("💾 Masukkan ke Database Cafe"):
                if nama_baru_cafe.strip() and sat_baru_cafe.strip():
                    conn = get_connection()
                    cursor = conn.cursor()
                    harga_satuan_cafe = float(hrg_baru_cafe / qty_baru_cafe)
                    cursor.execute("""
                        INSERT INTO bahan_baku (nama_bahan, harga_beli, jumlah_isi, satuan, harga_satuan, stok_cafe, harga_jual_internal)
                        VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT (nama_bahan) DO UPDATE SET stok_cafe = bahan_baku.stok_cafe + EXCLUDED.stok_cafe, harga_beli = EXCLUDED.harga_beli, harga_satuan = EXCLUDED.harga_satuan, harga_jual_internal = EXCLUDED.harga_jual_internal
                    """, (nama_baru_cafe.strip(), float(hrg_baru_cafe), float(qty_baru_cafe), sat_baru_cafe.strip(), harga_satuan_cafe, float(qty_baru_cafe), harga_satuan_cafe))
                    conn.commit()
                    conn.close()
                    st.toast(f"📦 {nama_baru_cafe} Masuk Stok Cafe!", icon="✅")
                    st.rerun()

    if role in ["Owner", "Admin"]:
        with tab_arsip:
            conn = get_connection()
            df_c = pd.read_sql_query('SELECT id as "ID", nama_resep as "Nama Menu", total_hpp_final as "HPP Pokok (Rp)", harga_jual as "Harga Jual (Rp)", tanggal_dibuat as "Waktu Pembuatan" FROM resep ORDER BY id DESC', conn)
            if df_c.empty: st.info("Belum ada resep.")
            else:
                edited = st.data_editor(df_c, use_container_width=True, disabled=["ID", "Nama Menu", "HPP Pokok (Rp)", "Waktu Pembuatan"])
                if st.button("💾 Simpan Perubahan Harga Jual", type="primary"):
                    cursor = conn.cursor()
                    for _, r in edited.iterrows():
                        cursor.execute("UPDATE resep SET harga_jual=%s WHERE id=%s", (float(r['Harga Jual (Rp)']), int(r['ID'])))
                    conn.commit()
                    st.toast("💰 Harga Diperbarui!", icon="✅")
                    st.rerun()
            conn.close()

# --- MENU 4: MENU OWNER ---
elif menu == "👥 Menu Owner":
    st.title("👥 Manajemen Otoritas Hak Akses User")
    conn = get_connection()
    df_users = pd.read_sql_query("SELECT username as \"ID User\", role as \"Tingkat Akses\" FROM users", conn)
    col1, col2 = st.columns([2, 1])
    with col1:
        st.subheader("📋 Daftar Pengguna Sistem")
        st.dataframe(df_users, use_container_width=True)
        hapus_user = st.selectbox("Hapus Akses Staf:", df_users['ID User'].tolist())
        if st.button("Hapus Akun Karyawan", type="primary", use_container_width=True):
            if hapus_user == "owner": st.error("Akses Owner Mutlak tidak boleh dihapus!")
            else:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM users WHERE username=%s", (hapus_user,))
                conn.commit()
                st.toast(f"🗑️ Akses {hapus_user} Dicabut!", icon="✅")
                st.rerun()
    with col2:
        st.subheader("➕ Daftarkan Akun Staf Baru")
        with st.form("tambah_staf", clear_on_submit=True):
            new_u = st.text_input("Username Baru:")
            new_p = st.text_input("Password Rahasia:", type="password")
            new_r = st.selectbox("Posisi Divisi:", ["Gudang Kios", "Karyawan", "Admin"])
            if st.form_submit_button("💾 Daftarkan Karyawan"):
                if new_u.strip() and new_p.strip():
                    cursor = conn.cursor()
                    try:
                        cursor.execute("INSERT INTO users VALUES (%s, %s, %s)", (new_u.strip().lower(), hash_password(new_p), new_r))
                        conn.commit()
                        st.toast(f"🎉 Akun {new_u} Berhasil Aktif!", icon="✅")
                        st.rerun()
                    except: st.error("ID tersebut sudah terdaftar!")
    conn.close()

render_footer()

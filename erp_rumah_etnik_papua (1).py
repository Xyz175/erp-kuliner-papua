import streamlit as st
import pandas as pd
import psycopg2
import hashlib
import io
import base64
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 0. TEMA WARNA (SESUAI LOGO RUMAH ETNIK PAPUA)
# ==========================================
C_PRIMARY = "#C57A2E"       # Oranye khas logo
C_PRIMARY_DARK = "#8F5620"
LOGO_FILE = "logo_rumah_etnik_papua.png"  # letakkan file ini 1 folder dengan script ini

# ==========================================
# 1. KONFIGURASI HALAMAN & DATABASE CLOUD
# ==========================================
st.set_page_config(
    page_title="ERP Rumah Etnik Papua",
    layout="wide",
    page_icon=LOGO_FILE if __import__("os").path.exists(LOGO_FILE) else "🏢"
)

# --- KREDENSIAL DATABASE ---
# PENTING (KEAMANAN): jangan menaruh username/password database langsung di kode.
# Simpan di file .streamlit/secrets.toml (lokal) atau menu "Secrets" (Streamlit Cloud):
#
#   SUPABASE_URI = "postgresql://user:password@host:port/db"
#
# Kode di bawah akan membaca dari st.secrets terlebih dahulu. Fallback hanya untuk
# menjaga aplikasi tetap berjalan sementara — SEGERA pindahkan ke secrets & GANTI
# password database ini karena sebelumnya sempat tertulis polos di source code.
_FALLBACK_URI = "postgresql://postgres.jskxygpvnbjgjjgsvwqf:ZeeyStore175@aws-0-ap-southeast-2.pooler.supabase.com:6543/postgres"
SUPABASE_URI = st.secrets.get("SUPABASE_URI", _FALLBACK_URI) if hasattr(st, "secrets") else _FALLBACK_URI


def get_connection():
    return psycopg2.connect(SUPABASE_URI)


def hash_password(raw_password: str) -> str:
    """Hash password dengan SHA-256 agar tidak tersimpan polos di database."""
    return hashlib.sha256(raw_password.strip().encode("utf-8")).hexdigest()


def init_db():
    """
    Fungsi ini dirancang khusus agar AMAN SAAT UPDATE KODE.
    Menggunakan 'IF NOT EXISTS' agar data lama di Supabase tidak terganggu atau terhapus.
    """
    conn = get_connection()
    conn.autocommit = True
    cursor = conn.cursor()

    # 1. Tabel Akses Akun (Owner, Admin, Karyawan, Gudang Kios)
    cursor.execute("CREATE TABLE IF NOT EXISTS users (username TEXT PRIMARY KEY, password TEXT, role TEXT)")
    # ON CONFLICT DO NOTHING menjamin password owner lama Anda tidak akan ter-reset saat update kode
    cursor.execute(
        "INSERT INTO users (username, password, role) VALUES ('owner', %s, 'Owner') ON CONFLICT DO NOTHING",
        (hash_password("owner123"),)
    )

    # Migrasi keamanan: hash ulang password lama yang masih tersimpan polos (bukan SHA-256 64 karakter)
    cursor.execute("SELECT username, password FROM users")
    for uname, pw in cursor.fetchall():
        if pw is None or len(pw) != 64:
            cursor.execute("UPDATE users SET password=%s WHERE username=%s", (hash_password(pw or ""), uname))

    # 2. Tabel Bahan Baku Utama (Gudang Pusat, Cafe, Taman)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bahan_baku (
            nama_bahan TEXT PRIMARY KEY, harga_beli REAL, jumlah_isi REAL, satuan TEXT, harga_satuan REAL,
            stok_gudang REAL DEFAULT 0, stok_cafe REAL DEFAULT 0, stok_taman REAL DEFAULT 0, harga_jual_internal REAL DEFAULT 0
        )
    """)

    # 3. Tabel Master Resep / Menu Finansial
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS resep (
            id SERIAL PRIMARY KEY, nama_resep TEXT UNIQUE, total_hpp_bahan REAL,
            total_operasional REAL, total_hpp_final REAL, harga_jual REAL DEFAULT 0, tanggal_dibuat TEXT
        )
    """)

    # 4. Tabel Detail Bahan Penyusun Menu
    cursor.execute("CREATE TABLE IF NOT EXISTS detail_resep (id SERIAL PRIMARY KEY, resep_id INTEGER, nama_bahan TEXT, jumlah_pakai REAL, satuan TEXT, subtotal_biaya REAL)")

    # 5. Tabel Log Riwayat Stok Logistik Pusat
    cursor.execute("CREATE TABLE IF NOT EXISTS riwayat_stok (id SERIAL PRIMARY KEY, tanggal TEXT, nama_bahan TEXT, jenis_transaksi TEXT, jumlah REAL, keterangan TEXT, laba_internal REAL DEFAULT 0, dari_gudang TEXT, ke_unit TEXT)")

    # 6. Tabel Mandiri Gudang Kios (Terpisah dari HPP Cafe)
    cursor.execute("CREATE TABLE IF NOT EXISTS stok_kios_item (nama_barang TEXT PRIMARY KEY, stok REAL DEFAULT 0, satuan TEXT)")
    cursor.execute("CREATE TABLE IF NOT EXISTS riwayat_kios (id SERIAL PRIMARY KEY, tanggal TEXT, nama_barang TEXT, jenis_transaksi TEXT, jumlah REAL, pic TEXT)")

    cursor.close()
    conn.close()


# Jalankan proteksi database
try:
    init_db()
except Exception as e:
    st.error(f"⚠️ Gagal sinkronisasi arsitektur DB: {e}")


# ==========================================
# 1B. HELPER EXPORT EXCEL YANG MENARIK (bukan CSV polos)
# ==========================================
def buat_excel_menarik(df: pd.DataFrame, title: str, sheet_name: str = "Data",
                        currency_cols=None, extra_info=None) -> bytes:
    """Mengubah DataFrame menjadi file Excel (.xlsx) berformat rapi: judul berwarna,
    header tebal, baris zebra, format Rupiah otomatis, dan lebar kolom otomatis."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(df.columns)
    n_cols = len(headers)
    last_col = get_column_letter(n_cols)
    thin = Side(style="thin", color="D8CBB8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    currency_cols = currency_cols or []

    row_cursor = 1
    ws.merge_cells(f"A{row_cursor}:{last_col}{row_cursor}")
    c = ws.cell(row=row_cursor, column=1, value=f"  {title}")
    c.font = Font(size=15, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="C57A2E", end_color="C57A2E", fill_type="solid")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_cursor].height = 30
    row_cursor += 1

    extra_info = extra_info or [f"Diekspor pada: {datetime.now().strftime('%d %B %Y, %H:%M')}"]
    for info in extra_info:
        ws.merge_cells(f"A{row_cursor}:{last_col}{row_cursor}")
        ic = ws.cell(row=row_cursor, column=1, value=f"  {info}")
        ic.font = Font(italic=True, size=9, color="7A6A57")
        row_cursor += 1
    row_cursor += 1

    header_row = row_cursor
    for col_idx, h in enumerate(headers, start=1):
        hc = ws.cell(row=header_row, column=col_idx, value=str(h))
        hc.font = Font(bold=True, color="FFFFFF", size=10)
        hc.fill = PatternFill(start_color="D98C3D", end_color="D98C3D", fill_type="solid")
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border = border
    row_cursor += 1

    col_max_len = [len(str(h)) for h in headers]
    for r_off, (_, row) in enumerate(df.iterrows()):
        r_idx = row_cursor + r_off
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if r_off % 2 == 1:
                cell.fill = PatternFill(start_color="FBF0E4", end_color="FBF0E4", fill_type="solid")
            if (c_idx - 1) in currency_cols and isinstance(val, (int, float)):
                cell.number_format = '"Rp" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            col_max_len[c_idx - 1] = max(col_max_len[c_idx - 1], len(str(val)))

    for col_idx, ml in enumerate(col_max_len, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(ml + 4, 12), 45)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def tombol_download_excel(df: pd.DataFrame, label: str, file_name: str, title: str,
                           currency_cols=None, key=None):
    if df is None or df.empty:
        return
    data = buat_excel_menarik(df, title=title, currency_cols=currency_cols)
    st.download_button(
        label, data=data, file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, key=key
    )


def get_daftar_bahan():
    try:
        conn = get_connection()
        df = pd.read_sql_query("SELECT * FROM bahan_baku ORDER BY nama_bahan ASC", conn)
        conn.close()
        return df
    except Exception as e:
        st.warning(f"Gagal memuat data bahan baku: {e}")
        return pd.DataFrame()


if 'racikan_sementara' not in st.session_state:
    st.session_state.racikan_sementara = pd.DataFrame(columns=["Nama Bahan", "Jumlah Pakai", "Satuan", "Subtotal"])
if 'sudah_login' not in st.session_state:
    st.session_state.sudah_login = False


# ==========================================
# 1C. GAYA VISUAL (CSS TEMA RUMAH ETNIK PAPUA)
# ==========================================
st.markdown(f"""
<style>
    .stApp {{ background-color: #FBF4EC; }}
    section[data-testid="stSidebar"] {{
        background-color: #FBF0E4;
        border-right: 3px solid {C_PRIMARY};
    }}
    h1, h2, h3 {{ color: {C_PRIMARY_DARK}; }}
    .stButton>button, .stDownloadButton>button {{
        background-color: {C_PRIMARY};
        color: white;
        border: none;
        border-radius: 8px;
        font-weight: 600;
    }}
    .stButton>button:hover, .stDownloadButton>button:hover {{
        background-color: {C_PRIMARY_DARK};
        color: white;
    }}
    div[data-testid="stMetric"] {{
        background-color: #FFFFFF;
        border: 1px solid #E8D3B8;
        border-radius: 10px;
        padding: 10px;
    }}
    .footer-zeey {{
        margin-top: 60px;
        padding: 22px 10px;
        text-align: center;
        border-top: 2px dashed {C_PRIMARY};
        color: #8F5620;
        font-size: 14px;
    }}
    .footer-zeey b {{ color: {C_PRIMARY_DARK}; }}
</style>
""", unsafe_allow_html=True)


def render_footer():
    st.markdown(f"""
    <div class="footer-zeey">
        🏛️ <b>ERP Rumah Etnik Papua</b> — Wisata Budaya Papua<br>
        Dibuat &amp; dikembangkan oleh <b>ZeeyStudio</b><br>
        📱 WhatsApp: <b>0813-5413-1178</b>
    </div>
    """, unsafe_allow_html=True)


# ==========================================
# 2. GERBANG UTAMA (SISTEM LOGIN) — DESAIN RUMAH ETNIK PAPUA
# ==========================================
if not st.session_state.sudah_login:
    import os
    col_hero, col_form = st.columns([1.1, 1])

    with col_hero:
        st.markdown(f"""
        <div style="background: linear-gradient(160deg, {C_PRIMARY} 0%, {C_PRIMARY_DARK} 100%);
                    border-radius: 18px; padding: 40px 30px; color: white; height: 430px;
                    display:flex; flex-direction:column; justify-content:center; align-items:center; text-align:center;">
            <h1 style="color:white; margin-bottom:4px;">Selamat Datang</h1>
            <p style="font-size:17px; opacity:0.95;">Sistem ERP Terpadu — Gudang, Cafe &amp; Kios</p>
            <p style="font-size:15px; opacity:0.85; margin-top:10px;">🏞️ Wisata Budaya Papua</p>
        </div>
        """, unsafe_allow_html=True)
        if os.path.exists(LOGO_FILE):
            st.image(LOGO_FILE, use_container_width=True)

    with col_form:
        st.markdown(f"""
        <div style="background:white; border:2px solid {C_PRIMARY}; border-radius:16px; padding:30px 30px 10px 30px;">
        """, unsafe_allow_html=True)
        st.subheader("🔐 Masuk ke Sistem")
        user_input = st.text_input("Username")
        pass_input = st.text_input("Password", type="password")
        if st.button("Masuk ke Sistem", type="primary", use_container_width=True):
            try:
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT role FROM users WHERE username=%s AND password=%s",
                    (user_input.strip().lower(), hash_password(pass_input))
                )
                res = cursor.fetchone()
                conn.close()
                if res:
                    st.session_state.sudah_login = True
                    st.session_state.username = user_input.strip().lower()
                    st.session_state.role = res[0]
                    st.rerun()
                else:
                    st.error("❌ Username atau Password salah!")
            except Exception as e:
                st.error(f"Gagal terhubung ke database: {e}")
        st.markdown("</div>", unsafe_allow_html=True)

    render_footer()
    st.stop()

# ==========================================
# 3. NAVIGASI BAR BERDASARKAN ROLE (LOGIS KETAT)
# ==========================================
import os
role = st.session_state.role
if os.path.exists(LOGO_FILE):
    st.sidebar.image(LOGO_FILE, use_container_width=True)
st.sidebar.title("🏢 KONTROL ERP")
st.sidebar.write(f"User: **{st.session_state.username}** ({role})")
if st.sidebar.button("🚪 Logout", use_container_width=True):
    st.session_state.sudah_login = False
    st.rerun()
st.sidebar.write("---")

if role == "Gudang Kios":
    menu_options = ["📦 Gudang"]
elif role == "Karyawan":
    menu_options = ["🍽️ HPP Cafe"]
elif role == "Admin":
    menu_options = ["📊 Dashboard Admin", "📦 Gudang", "🍽️ HPP Cafe"]
elif role == "Owner":
    menu_options = ["📊 Dashboard Admin", "📦 Gudang", "🍽️ HPP Cafe", "👥 Menu Owner"]
else:
    menu_options = []
    st.error("Role tidak dikenali. Hubungi Owner untuk memperbaiki akses akun Anda.")

menu = st.sidebar.radio("Navigasi:", menu_options) if menu_options else None

# ==========================================
# 4. IMPLEMENTASI OPERASIONAL
# ==========================================

# --- MENU 1: DASHBOARD ADMIN ---
if menu == "📊 Dashboard Admin":
    st.title("📊 Dashboard Analitik Finansial")
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT SUM(laba_internal) FROM riwayat_stok WHERE jenis_transaksi='Penjualan Internal'")
        total_laba_gudang = cursor.fetchone()[0] or 0.0

        cursor.execute("SELECT SUM(harga_jual - total_hpp_final) FROM resep WHERE harga_jual > 0")
        total_laba_cafe = cursor.fetchone()[0] or 0.0

        m1, m2 = st.columns(2)
        m1.metric("💰 Total Laba Internal Gudang Induk", f"Rp {total_laba_gudang:,.0f}")
        m2.metric("🍽️ Total Estimasi Laba Menu Cafe (per resep)", f"Rp {total_laba_cafe:,.0f}")

        st.write("---")
        st.subheader("📜 Riwayat Distribusi Logistik")
        df_log = pd.read_sql_query(
            "SELECT tanggal as \"Waktu\", nama_bahan as \"Barang\", jumlah as \"Qty\", ke_unit as \"Unit Tujuan\", "
            "keterangan as \"Omset\", laba_internal as \"Profit (Rp)\" FROM riwayat_stok ORDER BY id DESC", conn)
        st.dataframe(df_log, use_container_width=True)
        tombol_download_excel(
            df_log, "📥 Unduh Laporan Profit (Excel)", "Laporan_Profit_Gudang.xlsx",
            title="Laporan Distribusi & Profit Gudang Induk", currency_cols=[5]
        )
        conn.close()
    except Exception as e:
        st.warning(f"Menunggu sinkronisasi data... ({e})")

# --- MENU 2: GUDANG (TERINTEGRASI) ---
elif menu == "📦 Gudang":
    st.title("📦 Pusat Logistik & Manajemen Gudang")

    if role in ["Owner", "Admin"]:
        tab_induk, tab_transfer, tab_kios = st.tabs(["📥 Gudang Induk (Pusat)", "🚚 Transfer ke Cabang", "🏪 Gudang Kios (In/Out)"])
    else:
        tab_kios = st.container()

    if role in ["Owner", "Admin"]:
        with tab_induk:
            st.subheader("Manajemen Stok Gudang Induk")
            df_b = get_daftar_bahan()
            with st.form("form_pusat", clear_on_submit=True):
                cc1, cc2, cc3 = st.columns(3)
                with cc1: nama = st.text_input("Nama Komoditas Baru:")
                with cc2: hrg = st.number_input("Harga Beli Kulakan (Total Rp):", min_value=0.0)
                with cc3: qty = st.number_input("Kuantitas Isi Volume:", min_value=1.0)
                hrg_int = st.number_input("Harga Jual Internal ke Cabang (Rp / Satuan):", min_value=0.0)
                sat = st.text_input("Satuan Ukuran (gram, pcs, ml):")
                if st.form_submit_button("💾 Simpan ke Gudang Induk"):
                    if nama.strip() and sat.strip():
                        conn = get_connection()
                        cursor = conn.cursor()
                        harga_per_satuan = float(hrg / qty)
                        cursor.execute(
                            "INSERT INTO bahan_baku (nama_bahan, harga_beli, jumlah_isi, satuan, harga_satuan, stok_gudang, harga_jual_internal) "
                            "VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (nama_bahan) DO UPDATE SET "
                            "stok_gudang = bahan_baku.stok_gudang + EXCLUDED.stok_gudang, harga_jual_internal=EXCLUDED.harga_jual_internal",
                            (nama.strip(), float(hrg), float(qty), sat.strip(), harga_per_satuan, float(qty), float(hrg_int))
                        )
                        conn.commit()
                        conn.close()
                        st.toast(f"✅ {nama} Tersimpan di Gudang Induk!", icon="📥")
                        st.rerun()
                    else:
                        st.warning("Nama komoditas dan satuan wajib diisi.")
            st.write("---")
            if not df_b.empty:
                df_tampil = df_b[['nama_bahan', 'harga_satuan', 'harga_jual_internal', 'stok_gudang', 'stok_cafe', 'stok_taman', 'satuan']].copy()
                df_tampil.columns = ['Nama Barang', 'Modal Dasar', 'Jual Internal', 'STOK INDUK', 'STOK CAFE', 'STOK TAMAN', 'Satuan']
                st.dataframe(df_tampil, use_container_width=True)
                tombol_download_excel(
                    df_tampil, "📥 Unduh Data Bahan Baku (Excel)", "Data_Bahan_Baku.xlsx",
                    title="Data Bahan Baku Gudang Induk", currency_cols=[1, 2], key="dl_bahan_induk"
                )

        with tab_transfer:
            st.subheader("Transfer Logistik dari Gudang Induk ke Cabang")
            df_g = get_daftar_bahan()
            if df_g.empty:
                st.warning("Stok Gudang Induk kosong.")
            else:
                c1, c2, c3 = st.columns(3)
                with c1: brg = st.selectbox("Pilih Material:", df_g['nama_bahan'].tolist())
                data_b = df_g[df_g['nama_bahan'] == brg].iloc[0]
                stok_induk = data_b['stok_gudang']
                with c2: qty_tr = st.number_input(f"Kuantitas Kirim (Tersedia: {stok_induk}):", min_value=0.0, max_value=float(stok_induk) if stok_induk > 0 else 0.0)
                with c3: tuj = st.radio("Kirim Ke Cabang Mana?", ["Cafe", "Taman"])
                if st.button("🚀 Eksekusi Pengiriman", type="primary", use_container_width=True):
                    if qty_tr <= 0:
                        st.warning("Kuantitas kirim harus lebih dari 0.")
                    elif stok_induk >= qty_tr:
                        tgl = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        harga_acuan = data_b['harga_jual_internal'] if data_b['harga_jual_internal'] > 0 else data_b['harga_satuan']
                        laba = float((qty_tr * harga_acuan) - (qty_tr * data_b['harga_satuan']))
                        conn = get_connection()
                        cursor = conn.cursor()
                        cursor.execute("UPDATE bahan_baku SET stok_gudang = stok_gudang - %s WHERE nama_bahan = %s", (float(qty_tr), brg))
                        if tuj == "Cafe":
                            cursor.execute("UPDATE bahan_baku SET stok_cafe = COALESCE(stok_cafe, 0) + %s WHERE nama_bahan = %s", (float(qty_tr), brg))
                        elif tuj == "Taman":
                            cursor.execute("UPDATE bahan_baku SET stok_taman = COALESCE(stok_taman, 0) + %s WHERE nama_bahan = %s", (float(qty_tr), brg))
                        cursor.execute(
                            "INSERT INTO riwayat_stok (tanggal, nama_bahan, jenis_transaksi, jumlah, keterangan, laba_internal, dari_gudang, ke_unit) "
                            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                            (tgl, brg, 'Penjualan Internal', float(qty_tr), str(qty_tr * harga_acuan), laba, 'Gudang Etnik', tuj)
                        )
                        conn.commit()
                        conn.close()
                        st.toast(f"🚀 Berhasil Kirim ke {tuj}!", icon="🚚")
                        st.rerun()
                    else:
                        st.error("Stok Gudang Induk tidak mencukupi!")

    with tab_kios:
        st.subheader("Manajemen Mandiri Gudang Kios (Terpisah dari HPP)")
        conn = get_connection()
        df_kios = pd.read_sql_query("SELECT nama_barang, stok, satuan FROM stok_kios_item ORDER BY nama_barang ASC", conn)

        sub_tab_in, sub_tab_out, sub_tab_rep = st.tabs(["📥 Barang Masuk Kios", "📤 Barang Keluar Kios", "📊 Laporan Stok"])

        with sub_tab_in:
            pilihan = ["-- Input Barang Baru --"] + (df_kios['nama_barang'].tolist() if not df_kios.empty else [])
            barang_masuk = st.selectbox("Pilih Komoditas:", pilihan, key="kios_in_select")
            k1, k2, k3 = st.columns(3)
            with k1: nama_baru = st.text_input("Nama Produk", value="" if barang_masuk == "-- Input Barang Baru --" else barang_masuk, disabled=(barang_masuk != "-- Input Barang Baru --"), key="kios_in_name")
            with k2: qty_masuk = st.number_input("Kuantitas Masuk:", min_value=1.0, step=1.0, key="kios_in_qty")
            with k3: sat_baru = st.text_input("Satuan:", value="pcs" if barang_masuk == "-- Input Barang Baru --" else df_kios[df_kios['nama_barang'] == barang_masuk]['satuan'].values[0], key="kios_in_sat")
            if st.button("📥 Validasi Masuk Kios", type="primary"):
                if nama_baru.strip():
                    cursor = conn.cursor()
                    tgl = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    cursor.execute(
                        "INSERT INTO stok_kios_item (nama_barang, stok, satuan) VALUES (%s, %s, %s) "
                        "ON CONFLICT (nama_barang) DO UPDATE SET stok = stok_kios_item.stok + EXCLUDED.stok",
                        (nama_baru.strip(), float(qty_masuk), sat_baru.strip())
                    )
                    cursor.execute(
                        "INSERT INTO riwayat_kios (tanggal, nama_barang, jenis_transaksi, jumlah, pic) VALUES (%s, %s, %s, %s, %s)",
                        (tgl, nama_baru.strip(), 'Masuk', float(qty_masuk), st.session_state.username)
                    )
                    conn.commit()
                    st.toast("📥 Stok Kios Berhasil Diupdate!", icon="🏪")
                    st.rerun()
                else:
                    st.warning("Nama produk wajib diisi.")

        with sub_tab_out:
            if df_kios.empty:
                st.warning("Stok Kios Masih Kosong!")
            else:
                ko1, ko2 = st.columns(2)
                with ko1: barang_keluar = st.selectbox("Pilih Produk Keluar:", df_kios['nama_barang'].tolist(), key="kios_out_select")
                stok_sisa = df_kios[df_kios['nama_barang'] == barang_keluar]['stok'].values[0]
                with ko2: qty_keluar = st.number_input(f"Kuantitas Keluar (Tersedia: {stok_sisa}):", min_value=0.0, max_value=float(stok_sisa) if stok_sisa > 0 else 0.0, step=1.0, key="kios_out_qty")
                if st.button("📤 Validasi Keluar Kios", type="primary"):
                    if qty_keluar <= 0:
                        st.warning("Kuantitas keluar harus lebih dari 0.")
                    else:
                        cursor = conn.cursor()
                        tgl = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        cursor.execute("UPDATE stok_kios_item SET stok = stok - %s WHERE nama_barang = %s", (float(qty_keluar), barang_keluar))
                        cursor.execute(
                            "INSERT INTO riwayat_kios (tanggal, nama_barang, jenis_transaksi, jumlah, pic) VALUES (%s, %s, %s, %s, %s)",
                            (tgl, barang_keluar, 'Keluar', float(qty_keluar), st.session_state.username)
                        )
                        conn.commit()
                        st.toast("📤 Mutasi Keluar Kios Tercatat!", icon="🛒")
                        st.rerun()

        with sub_tab_rep:
            st.dataframe(df_kios, use_container_width=True)
            tombol_download_excel(df_kios, "📥 Unduh Stok Kios (Excel)", "Stok_Kios.xlsx", title="Laporan Stok Gudang Kios", key="dl_stok_kios")
            st.write("---")
            df_log_kios = pd.read_sql_query(
                "SELECT tanggal as \"Waktu\", nama_barang as \"Barang\", jenis_transaksi as \"Status\", jumlah as \"Qty\", "
                "pic as \"Operator\" FROM riwayat_kios ORDER BY id DESC", conn)
            st.dataframe(df_log_kios, use_container_width=True)
            tombol_download_excel(df_log_kios, "📥 Unduh Riwayat Kios (Excel)", "Riwayat_Kios.xlsx", title="Riwayat Transaksi Gudang Kios", key="dl_riwayat_kios")
        conn.close()

# --- MENU 3: HPP CAFE ---
elif menu == "🍽️ HPP Cafe":
    st.title("🍽️ Pusat Operasional & Manajemen HPP Cafe")

    if role in ["Owner", "Admin"]:
        tab_dapur, tab_tambah_bahan, tab_arsip = st.tabs(["🍳 Input Pemakaian Dapur", "➕ Tambah Bahan Cafe", "📜 Arsip & Harga Jual"])
    else:
        tab_dapur, tab_tambah_bahan = st.tabs(["🍳 Input Pemakaian Dapur", "➕ Tambah Bahan Cafe"])

    # TAB 1: RACIK MENU & POTONG STOK
    with tab_dapur:
        st.subheader("🍳 Meracik Menu & Potong Stok")
        df_b = get_daftar_bahan()
        if df_b.empty:
            st.warning("Belum ada bahan. Silakan isi di tab 'Tambah Bahan Cafe' di sebelah.")
        else:
            c1, c2, c3 = st.columns([2, 1, 1])
            with c1:
                cari = st.text_input("🔍 Cari bahan baku Cafe:")
                df_f = df_b[df_b['nama_bahan'].str.lower().str.contains(cari.lower())] if cari else df_b
                pilih_b = st.selectbox("Pilih item:", df_f['nama_bahan'].tolist() if not df_f.empty else ["Kosong"])
            with c2: qty = st.number_input("Volume / Jumlah Pakai:", min_value=0.0, step=1.0, key="cafe_qty_input")
            with c3:
                st.write(""); st.write("")
                if st.button("➕ Tambah ke Draf Menu", use_container_width=True):
                    if pilih_b != "Kosong" and qty > 0:
                        row = df_b[df_b['nama_bahan'] == pilih_b].iloc[0]
                        h_modal = row['harga_jual_internal'] if row['harga_jual_internal'] > 0 else row['harga_satuan']
                        sub = float(qty * h_modal)
                        df_skrg = st.session_state.racikan_sementara
                        if pilih_b in df_skrg['Nama Bahan'].values:
                            df_skrg.loc[df_skrg['Nama Bahan'] == pilih_b, ["Jumlah Pakai", "Subtotal"]] = [float(qty), sub]
                        else:
                            tambah = pd.DataFrame([[str(pilih_b), float(qty), str(row['satuan']), sub]], columns=df_skrg.columns)
                            st.session_state.racikan_sementara = pd.concat([df_skrg, tambah], ignore_index=True)
                        st.rerun()
                    else:
                        st.warning("Pilih bahan dan isi jumlah pakai lebih dari 0.")
        st.write("---")
        if not st.session_state.racikan_sementara.empty:
            st.session_state.racikan_sementara = st.data_editor(st.session_state.racikan_sementara, use_container_width=True)
            tot_bahan = float(st.session_state.racikan_sementara['Subtotal'].sum())

            st.subheader("⚡ Komponen Operasional Dapur")
            co1, co2 = st.columns(2)
            with co1: waktu = st.number_input("Waktu Meracik (Menit):", min_value=0.0, step=5.0)
            with co2: gas = st.number_input("Estimasi Gas LPG (Kg):", min_value=0.0, format="%.3f")

            biaya_ops = float(((10000 / 60) * waktu) + ((3000 / 60) * waktu) + (gas * 18000))
            tot_final = float(tot_bahan + biaya_ops)

            st.info(f"📊 Finansial Resep: Modal Bahan: Rp {tot_bahan:,.0f} | Operasional: Rp {biaya_ops:,.0f} | HPP Pokok: Rp {tot_final:,.2f}")

            st.subheader("💰 Estimasi Harga Jual & Laba")
            hj1, hj2 = st.columns(2)
            with hj1:
                harga_jual_draf = st.number_input("Rencana Harga Jual Menu (Rp):", min_value=0.0, step=1000.0, key="harga_jual_draf")
            with hj2:
                laba_draf = harga_jual_draf - tot_final
                margin_draf = (laba_draf / harga_jual_draf * 100) if harga_jual_draf > 0 else 0.0
                warna_laba = "green" if laba_draf >= 0 else "red"
                st.markdown(f"**Estimasi Laba:** :{warna_laba}[Rp {laba_draf:,.2f} ({margin_draf:.1f}%)]")

            cx1, cx2 = st.columns(2)
            with cx1: nama_menu = st.text_input("Nama Menu Kuliner Baru:")
            with cx2:
                st.write(""); st.write("")
                if st.button("💾 Kunci Resep & Potong Stok", type="primary", use_container_width=True):
                    if nama_menu.strip():
                        conn = get_connection()
                        cursor = conn.cursor()
                        tgl = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        cursor.execute(
                            "INSERT INTO resep (nama_resep, total_hpp_bahan, total_operasional, total_hpp_final, harga_jual, tanggal_dibuat) "
                            "VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (nama_resep) DO UPDATE SET "
                            "total_hpp_bahan=EXCLUDED.total_hpp_bahan, total_operasional=EXCLUDED.total_operasional, "
                            "total_hpp_final=EXCLUDED.total_hpp_final, harga_jual=EXCLUDED.harga_jual",
                            (nama_menu.strip(), tot_bahan, biaya_ops, tot_final, harga_jual_draf, tgl)
                        )
                        cursor.execute("SELECT id FROM resep WHERE nama_resep=%s", (nama_menu.strip(),))
                        r_id = cursor.fetchone()[0]
                        cursor.execute("DELETE FROM detail_resep WHERE resep_id=%s", (r_id,))

                        for _, r in st.session_state.racikan_sementara.iterrows():
                            nama_bhn = str(r['Nama Bahan'])
                            jml_pakai = float(r['Jumlah Pakai'])
                            cursor.execute(
                                "INSERT INTO detail_resep (resep_id, nama_bahan, jumlah_pakai, satuan, subtotal_biaya) VALUES (%s,%s,%s,%s,%s)",
                                (r_id, nama_bhn, jml_pakai, str(r['Satuan']), float(r['Subtotal']))
                            )
                            cursor.execute("UPDATE bahan_baku SET stok_cafe = COALESCE(stok_cafe,0) - %s WHERE nama_bahan = %s", (jml_pakai, nama_bhn))

                        conn.commit()
                        conn.close()
                        st.session_state.racikan_sementara = pd.DataFrame(columns=["Nama Bahan", "Jumlah Pakai", "Satuan", "Subtotal"])
                        st.toast("🎉 Resep Menu Berhasil Disimpan!", icon="🍳")
                        st.rerun()
                    else:
                        st.warning("Nama menu wajib diisi.")

    # TAB 2: TAMBAH BAHAN CAFE MANDIRI
    with tab_tambah_bahan:
        st.subheader("➕ Tambah Bahan Baku Belanjaan Baru")
        with st.form("form_tambah_bahan_cafe", clear_on_submit=True):
            col_b1, col_b2 = st.columns(2)
            with col_b1: nama_baru_cafe = st.text_input("Nama Bahan Baku:")
            with col_b2: hrg_baru_cafe = st.number_input("Harga Beli (Total Rp):", min_value=0.0)
            col_b3, col_b4 = st.columns(2)
            with col_b3: qty_baru_cafe = st.number_input("Jumlah / Isi Berat:", min_value=1.0)
            with col_b4: sat_baru_cafe = st.text_input("Satuan (kg, pcs, ml, dll):")

            if st.form_submit_button("💾 Masukkan ke Database Cafe"):
                if nama_baru_cafe.strip() and sat_baru_cafe.strip():
                    conn = get_connection()
                    cursor = conn.cursor()
                    harga_satuan_cafe = float(hrg_baru_cafe / qty_baru_cafe)
                    cursor.execute("""
                        INSERT INTO bahan_baku (nama_bahan, harga_beli, jumlah_isi, satuan, harga_satuan, stok_cafe, harga_jual_internal)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (nama_bahan)
                        DO UPDATE SET stok_cafe = bahan_baku.stok_cafe + EXCLUDED.stok_cafe,
                                      harga_beli = EXCLUDED.harga_beli,
                                      harga_satuan = EXCLUDED.harga_satuan,
                                      harga_jual_internal = EXCLUDED.harga_jual_internal
                    """, (nama_baru_cafe.strip(), float(hrg_baru_cafe), float(qty_baru_cafe), sat_baru_cafe.strip(), harga_satuan_cafe, float(qty_baru_cafe), harga_satuan_cafe))
                    conn.commit()
                    conn.close()
                    st.toast(f"📦 {nama_baru_cafe} Sukses Ditambahkan ke Stok Cafe!", icon="✅")
                    st.rerun()
                else:
                    st.warning("Nama bahan dan satuan wajib diisi.")

    # TAB 3: ARSIP CAFE (HANYA OWNER/ADMIN)
    if role in ["Owner", "Admin"]:
        with tab_arsip:
            st.subheader("📜 Arsip Dokumen HPP & Manajemen Harga Jual Cafe")
            conn = get_connection()
            df_c = pd.read_sql_query(
                'SELECT id as "ID", nama_resep as "Nama Menu", total_hpp_final as "HPP Pokok (Rp)", '
                'harga_jual as "Harga Jual (Rp)", tanggal_dibuat as "Waktu Pembuatan" FROM resep ORDER BY id DESC', conn)

            if df_c.empty:
                st.info("Belum ada resep terdokumentasi.")
            else:
                df_c["Laba (Rp)"] = df_c["Harga Jual (Rp)"] - df_c["HPP Pokok (Rp)"]
                df_c["Margin (%)"] = df_c.apply(
                    lambda r: (r["Laba (Rp)"] / r["Harga Jual (Rp)"] * 100) if r["Harga Jual (Rp)"] > 0 else 0.0, axis=1
                ).round(1)

                m1, m2, m3 = st.columns(3)
                m1.metric("🍽️ Total Menu", len(df_c))
                m2.metric("💰 Total Estimasi Laba", f"Rp {df_c['Laba (Rp)'].sum():,.0f}")
                m3.metric("📈 Rata-rata Margin", f"{df_c['Margin (%)'].mean():.1f}%")

                edited = st.data_editor(
                    df_c, use_container_width=True,
                    disabled=["ID", "Nama Menu", "HPP Pokok (Rp)", "Waktu Pembuatan", "Laba (Rp)", "Margin (%)"]
                )
                c_as1, c_as2 = st.columns(2)
                with c_as1:
                    if st.button("💾 Simpan Perubahan Harga Jual", type="primary", use_container_width=True):
                        cursor = conn.cursor()
                        for _, r in edited.iterrows():
                            cursor.execute("UPDATE resep SET harga_jual=%s WHERE id=%s", (float(r['Harga Jual (Rp)']), int(r['ID'])))
                        conn.commit()
                        st.toast("💰 Harga Jual Berhasil Diperbarui!", icon="✅")
                        st.rerun()
                with c_as2:
                    tombol_download_excel(
                        df_c, "📥 Unduh Laporan HPP & Laba Cafe (Excel)", "Arsip_HPP_Cafe.xlsx",
                        title="Arsip HPP, Harga Jual & Laba Menu Cafe", currency_cols=[2, 3, 4], key="dl_arsip_cafe"
                    )

                st.write("---")
                pilih_del = st.selectbox("Pilih Menu untuk Dihapus Permanen:", df_c['Nama Menu'].tolist())
                if st.button("🗑️ Hapus Menu Permanen", use_container_width=True):
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM detail_resep WHERE resep_id = (SELECT id FROM resep WHERE nama_resep=%s)", (pilih_del,))
                    cursor.execute("DELETE FROM resep WHERE nama_resep=%s", (pilih_del,))
                    conn.commit()
                    st.toast(f"🗑️ Menu {pilih_del} Berhasil Dihapus!", icon="⚠️")
                    st.rerun()
            conn.close()

# --- MENU 4: MENU OWNER (OTORITAS TINGGI) ---
elif menu == "👥 Menu Owner":
    st.title("👥 Manajemen Otoritas Hak Akses User")
    conn = get_connection()
    df_users = pd.read_sql_query("SELECT username as \"ID User\", role as \"Tingkat Akses\" FROM users", conn)
    col1, col2 = st.columns([2, 1])
    with col1:
        st.subheader("📋 Daftar Pengguna Sistem")
        st.dataframe(df_users, use_container_width=True)
        hapus_user = st.selectbox("Hapus Akses Staf:", df_users['ID User'].tolist())
        if st.button("Hapus Akun Karyawan", type="primary", use_container_width=True):
            if hapus_user == "owner":
                st.error("Akses Owner Mutlak tidak boleh dihapus!")
            else:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM users WHERE username=%s", (hapus_user,))
                conn.commit()
                st.toast(f"🗑️ Akses {hapus_user} Dicabut!", icon="✅")
                st.rerun()
    with col2:
        st.subheader("➕ Daftarkan Akun Staf Baru")
        with st.form("tambah_staf", clear_on_submit=True):
            new_u = st.text_input("Username Baru:")
            new_p = st.text_input("Password Rahasia:", type="password")
            # Catatan keamanan: role "Owner" sengaja TIDAK bisa dipilih di sini agar
            # tidak ada duplikasi otoritas tertinggi tanpa sepengetahuan pemilik asli.
            new_r = st.selectbox("Posisi Divisi:", ["Gudang Kios", "Karyawan", "Admin"])
            if st.form_submit_button("💾 Daftarkan Karyawan"):
                if new_u.strip() and new_p.strip():
                    cursor = conn.cursor()
                    try:
                        cursor.execute(
                            "INSERT INTO users VALUES (%s, %s, %s)",
                            (new_u.strip().lower(), hash_password(new_p), new_r)
                        )
                        conn.commit()
                        st.toast(f"🎉 Akun {new_u} Berhasil Aktif!", icon="✅")
                        st.rerun()
                    except Exception:
                        st.error("ID tersebut sudah terdaftar!")
                else:
                    st.warning("Username dan password wajib diisi.")
    conn.close()

render_footer()
